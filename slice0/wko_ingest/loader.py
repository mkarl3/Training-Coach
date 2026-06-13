"""Load WKO5 .xlsx exports into SQLite per the approved Slice-0 schema.

Field mapping is keyed on the row-2 header NAMES (not column index), so year-to-year
column drift (PMC gaining HRV/RHR/sleep, the partial/weekly files dropping 2-Hour Power)
is handled transparently.
"""
import datetime
import glob
import os
import sqlite3

from . import parse

HERE = os.path.dirname(os.path.abspath(__file__))
SCHEMA_PATH = os.path.join(HERE, "schema.sql")

# Header NAME -> (workout column, parser). Keyed on row-2 text.
TH_FIELDS = {
    "Activity Type": ("activity_type", parse.parse_str),
    "Total Duration": ("duration_sec", parse.parse_duration_sec),
    "Total Distance": ("distance_mi", parse.parse_float),
    "TSS": ("tss", parse.parse_float),
    "Work": ("work_kj", parse.parse_float),
    "Normalized Power": ("np_w", parse.parse_float),
    "Avg Heart Rate": ("avg_hr_bpm", parse.parse_float),
    "Max Heart Rate": ("max_hr_bpm", parse.parse_float),
    "Cadence": ("cadence_rpm", parse.parse_float),
    "IF": ("if_", parse.parse_float),
    "EF": ("ef", parse.parse_float),
    "VI": ("vi", parse.parse_float),
    "5 Sec Power": ("p5s_w", parse.parse_float),
    "1 Min Power": ("p1min_w", parse.parse_float),
    "5 Min Power": ("p5min_w", parse.parse_float),
    "10 Min Power": ("p10min_w", parse.parse_float),
    "20 Min Power": ("p20min_w", parse.parse_float),
    "1 Hour Power": ("p1hr_w", parse.parse_float),
    "2 Hour Power": ("p2hr_w", parse.parse_float),
    "RPE": ("rpe", parse.parse_float),
    "Feeling": ("feeling", parse.parse_float),
    "Anaerobic Training Impact Score": ("anaerobic_tis", parse.parse_float),
    "Aerobic Training Impact Score": ("aerobic_tis", parse.parse_float),
    "pwHr": ("pwhr_pct", parse.parse_float),
}

# PMC midnight-row metrics.
PMC_METRIC_FIELDS = {
    "ATL": ("atl", parse.parse_float),
    "CTL": ("ctl", parse.parse_float),
    "TSB": ("tsb", parse.parse_float),
    "mFTP": ("mftp_w", parse.parse_float),
    "FRC": ("frc_kj", parse.parse_float),
    "PMax": ("pmax_w", parse.parse_float),
    "TTE": ("tte_sec", parse.parse_duration_sec),
}
# PMC wellness fields (may appear on midnight or intraday rows; coalesced per date).
PMC_WELLNESS_FIELDS = {
    "Weight": ("weight_lb", parse.parse_float),
    "Fat%": ("fat_pct", parse.parse_float),
    "Sickness": ("sickness", parse.parse_str),
    "7d Avg HRV": ("hrv_7d_avg_ms", parse.parse_float),
    "Daily HRV": ("hrv_daily_ms", parse.parse_float),
    "RHR": ("rhr_bpm", parse.parse_float),
    "Sleep Hours": ("sleep_total_sec", lambda v: parse.parse_duration_sec(v, "hm")),
    "Time in Deep Sleep": ("sleep_deep_sec", lambda v: parse.parse_duration_sec(v, "hm")),
    "Time in Light Sleep": ("sleep_light_sec", lambda v: parse.parse_duration_sec(v, "hm")),
    "Time In Rem Sleep": ("sleep_rem_sec", lambda v: parse.parse_duration_sec(v, "hm")),
    "Time Awake": ("sleep_awake_sec", lambda v: parse.parse_duration_sec(v, "hm")),
}
TIZ_FIELDS = {
    "TiZ ClassicPower Z1": ("tiz_pwr_z1_sec", parse.parse_duration_sec),
    "TiZ ClassicPower Z2": ("tiz_pwr_z2_sec", parse.parse_duration_sec),
    "TiZ ClassicPower Z3": ("tiz_pwr_z3_sec", parse.parse_duration_sec),
    "TiZ ClassicPower Z4": ("tiz_pwr_z4_sec", parse.parse_duration_sec),
    "TiZ ClassicPower Z5": ("tiz_pwr_z5_sec", parse.parse_duration_sec),
    "TiZ ClassicPower Z6": ("tiz_pwr_z6_sec", parse.parse_duration_sec),
    "TiZ Classic HR Z1": ("tiz_hr_z1_sec", parse.parse_duration_sec),
    "TiZ Classic HR Z2": ("tiz_hr_z2_sec", parse.parse_duration_sec),
    "TiZ Classic HR Z3": ("tiz_hr_z3_sec", parse.parse_duration_sec),
    "TiZ Classic HR Z4": ("tiz_hr_z4_sec", parse.parse_duration_sec),
    "TiZ Classic HR Z5": ("tiz_hr_z5_sec", parse.parse_duration_sec),
}


def _is_cycling(activity_type):
    if not activity_type:
        return False
    a = activity_type.lower()
    return ("bike" in a) or ("cycl" in a) or ("ride" in a)


def header_map(ws):
    """Map row-2 header name -> 0-based column index, and find the date column.

    The date column is whichever column's row-3 unit text starts with 'date'.
    """
    rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
    names = rows[0] if rows else ()
    units = rows[1] if len(rows) > 1 else ()
    name_to_idx = {}
    for idx, n in enumerate(names):
        if isinstance(n, str) and n.strip():
            name_to_idx[n.strip()] = idx
    date_idx = None
    for idx, u in enumerate(units):
        if isinstance(u, str) and u.strip().lower().startswith("date"):
            date_idx = idx
            break
    if date_idx is None:
        date_idx = 1  # col B fallback
    return name_to_idx, date_idx


def _data_rows(ws, date_idx):
    for row in ws.iter_rows(min_row=4, values_only=True):
        if date_idx >= len(row):
            continue
        if row[date_idx] is None:
            continue
        yield row


def read_training_history(ws, source_file):
    """Yield per-workout dicts from a Training History sheet."""
    names, date_idx = header_map(ws)
    out = []
    for row in _data_rows(ws, date_idx):
        started = parse.parse_datetime(row[date_idx])
        d = parse.parse_date(row[date_idx])
        if d is None:
            continue
        rec = {"date": d, "started_at": started, "source_file": source_file}
        for name, (col, fn) in TH_FIELDS.items():
            idx = names.get(name)
            rec[col] = fn(row[idx]) if idx is not None and idx < len(row) else None
        rec["is_cycling"] = 1 if _is_cycling(rec.get("activity_type")) else 0
        out.append(rec)
    return out


def read_pmc(ws):
    """Return {date: {metric/wellness fields}} merging the dual rows per date.

    Midnight (00:00) row -> metric fields. Wellness fields are coalesced (first
    non-null) across all rows for that date. Returns (per_date, conflicts).
    """
    names, date_idx = header_map(ws)
    per_date = {}
    conflicts = []
    for row in _data_rows(ws, date_idx):
        raw = row[date_idx]
        d = parse.parse_date(raw)
        if d is None:
            continue
        is_midnight = isinstance(raw, datetime.datetime) and (raw.hour, raw.minute, raw.second) == (0, 0, 0)
        bucket = per_date.setdefault(d, {})
        if is_midnight:
            for name, (col, fn) in PMC_METRIC_FIELDS.items():
                idx = names.get(name)
                if idx is not None and idx < len(row):
                    val = fn(row[idx])
                    if val is not None:
                        bucket[col] = val
        # Wellness can sit on either row; take first non-null, note conflicts.
        for name, (col, fn) in PMC_WELLNESS_FIELDS.items():
            idx = names.get(name)
            if idx is None or idx >= len(row):
                continue
            val = fn(row[idx])
            if val is None:
                continue
            if col in bucket and bucket[col] != val:
                conflicts.append((d, col, bucket[col], val))
            else:
                bucket.setdefault(col, val)
    return per_date, conflicts


def read_tiz(ws):
    """Return {date: {tiz_* fields}} from a Daily TiZ sheet."""
    names, date_idx = header_map(ws)
    per_date = {}
    for row in _data_rows(ws, date_idx):
        d = parse.parse_date(row[date_idx])
        if d is None:
            continue
        bucket = per_date.setdefault(d, {})
        for name, (col, fn) in TIZ_FIELDS.items():
            idx = names.get(name)
            if idx is not None and idx < len(row):
                val = fn(row[idx])
                if val is not None:
                    bucket[col] = val
    return per_date


def dedup_workouts(workouts):
    """Dedup by (started_at, activity_type, duration_sec). Keeps LAST seen — callers pass
    workouts in ascending precedence order, so the newer file's version wins."""
    by_key = {}
    for w in workouts:
        by_key[(w["started_at"], w.get("activity_type"), w.get("duration_sec"))] = w
    return list(by_key.values())


def _classify(basename):
    if basename.startswith("Daily TiZ"):
        return "TiZ"
    if basename.startswith("PMC"):
        return "PMC"
    if basename.startswith("Training History"):
        return "TH"
    if basename.startswith("Week"):
        return "Week"
    return "?"


def _date_range(dates):
    ds = [d for d in dates if d]
    return (min(ds), max(ds)) if ds else (None, None)


def _all_days(dmin, dmax):
    a = datetime.date.fromisoformat(dmin)
    b = datetime.date.fromisoformat(dmax)
    out = []
    cur = a
    while cur <= b:
        out.append(cur.isoformat())
        cur += datetime.timedelta(days=1)
    return out


def build_database(db_path, exports_dir, loaded_at=None):
    """Parse all exports and populate `db_path`. Returns a load-summary dict."""
    import openpyxl

    if loaded_at is None:
        loaded_at = "unknown"  # caller stamps real time; scripts avoid Date.now-style calls
    # PRECEDENCE: a date covered by multiple exports resolves to the NEWER file (by mtime);
    # on an mtime tie, a full-year file beats a weekly snapshot. We process lowest-precedence
    # first so the winner is applied LAST — PMC/TiZ via dict-update, workouts via keep-last
    # dedup. Weekly files are first-class: all three sheets are ingested like the yearly ones.
    # (Going forward, weekly files cover new, non-overlapping weeks; precedence only matters
    # where exports overlap, e.g. a weekly file replacing a yearly file's PMC projection rows.)
    def _precedence(f):
        is_week = _classify(os.path.basename(f)) == "Week"
        return (os.path.getmtime(f), 0 if is_week else 1)   # weekly before yearly on tie
    files = sorted(glob.glob(os.path.join(exports_dir, "*.xlsx")), key=_precedence)

    meta_rows = []
    workouts = []
    pmc_by_date = {}
    tiz_by_date = {}
    pmc_conflicts = []

    for f in files:
        base = os.path.basename(f)
        fam = _classify(base)
        wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
        try:
            for ws in wb.worksheets:
                sheet_fam = fam
                if fam == "Week":
                    title = ws.title.lower()
                    sheet_fam = ("TH" if "training" in title else
                                 "PMC" if "pmc" in title else
                                 "TiZ" if "tiz" in title else fam)

                if sheet_fam == "TH":
                    recs = read_training_history(ws, base)
                    workouts.extend(recs)
                    dmin, dmax = _date_range([r["date"] for r in recs])
                    meta_rows.append((base, ws.title, sheet_fam, "loaded", len(recs),
                                      len(recs), 0, dmin, dmax))

                elif sheet_fam == "PMC":
                    per_date, conflicts = read_pmc(ws)
                    pmc_conflicts.extend((base,) + c for c in conflicts)
                    for d, vals in per_date.items():
                        pmc_by_date.setdefault(d, {}).update(vals)   # newer applied last -> wins
                    dmin, dmax = _date_range(list(per_date))
                    meta_rows.append((base, ws.title, sheet_fam, "loaded", len(per_date),
                                      len(per_date), 0, dmin, dmax))

                elif sheet_fam == "TiZ":
                    per_date = read_tiz(ws)
                    for d, vals in per_date.items():
                        tiz_by_date.setdefault(d, {}).update(vals)
                    dmin, dmax = _date_range(list(per_date))
                    meta_rows.append((base, ws.title, sheet_fam, "loaded", len(per_date),
                                      len(per_date), 0, dmin, dmax))
        finally:
            wb.close()

    workouts = dedup_workouts(workouts)   # keep-last == newer file wins (processed last)

    # --- horizon: last ACTUAL ride day. Wellness-only days do not extend it. ---
    ride_dates = [w["date"] for w in workouts if w["is_cycling"]]
    horizon = max(ride_dates) if ride_dates else None

    # --- date spine: full calendar span across every loaded source (incl. PMC future). ---
    all_dates = ([w["date"] for w in workouts] + list(pmc_by_date) + list(tiz_by_date))
    dmin, dmax = _date_range(all_dates)
    spine = _all_days(dmin, dmax)

    # --- per-day workout aggregates ---
    agg = {}
    for w in workouts:
        a = agg.setdefault(w["date"], {"n": 0, "ride": 0, "tss": 0.0, "dur": 0,
                                       "dist": 0.0, "work": 0.0, "if_num": 0.0, "if_den": 0})
        a["n"] += 1
        if w["is_cycling"]:
            a["ride"] = 1
        a["tss"] += w["tss"] or 0.0
        a["dur"] += w["duration_sec"] or 0
        a["dist"] += w["distance_mi"] or 0.0
        a["work"] += w["work_kj"] or 0.0
        if w["if_"] is not None and w["duration_sec"]:
            a["if_num"] += w["if_"] * w["duration_sec"]
            a["if_den"] += w["duration_sec"]

    daily_rows = []
    for d in spine:
        projected = 1 if (horizon is not None and d > horizon) else 0
        a = agg.get(d)
        if projected:
            tss_sum = dur = dist = work = if_daily = None
            n = a["n"] if a else 0
            ride = a["ride"] if a else 0
        else:
            if a:
                tss_sum = a["tss"]
                dur = a["dur"]
                dist = a["dist"]
                work = a["work"]
                if_daily = (a["if_num"] / a["if_den"]) if a["if_den"] else None
                n = a["n"]
                ride = a["ride"]
            else:
                # actual no-ride / rest day: trained zero, NOT unknown
                tss_sum, dur, dist, work, if_daily, n, ride = 0.0, 0, 0.0, 0.0, None, 0, 0
        rec = {
            "date": d, "year": int(d[:4]), "is_projected": projected,
            "has_ride": ride, "num_workouts": n,
            "tss_sum": tss_sum, "duration_sec": dur, "distance_mi": dist,
            "work_kj": work, "if_daily": if_daily, "data_flags": None,
        }
        for col in ("atl", "ctl", "tsb", "mftp_w", "frc_kj", "pmax_w", "tte_sec"):
            rec[col] = pmc_by_date.get(d, {}).get(col)
        for col in ("weight_lb", "fat_pct", "sickness", "hrv_7d_avg_ms", "hrv_daily_ms",
                    "rhr_bpm", "sleep_total_sec", "sleep_deep_sec", "sleep_light_sec",
                    "sleep_rem_sec", "sleep_awake_sec"):
            rec[col] = pmc_by_date.get(d, {}).get(col)
        for _, (col, _fn) in TIZ_FIELDS.items():
            rec[col] = tiz_by_date.get(d, {}).get(col)
        daily_rows.append(rec)

    # --- write SQLite ---
    if os.path.exists(db_path):
        os.remove(db_path)
    conn = sqlite3.connect(db_path)
    try:
        with open(SCHEMA_PATH, "r", encoding="utf-8") as fh:
            conn.executescript(fh.read())
        _insert_workouts(conn, workouts)
        _insert_daily(conn, daily_rows)
        conn.executemany(
            "INSERT INTO ingest_meta (source_file,sheet,family,role,rows_read,rows_loaded,"
            "rows_rejected,date_min,date_max,loaded_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            [m + (loaded_at,) for m in meta_rows],
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "db_path": db_path,
        "workouts": len(workouts),
        "daily_rows": len(daily_rows),
        "horizon": horizon,
        "date_min": dmin,
        "date_max": dmax,
        "pmc_conflicts": pmc_conflicts,
        "files": len(files),
    }


_WORKOUT_COLS = [
    "date", "started_at", "activity_type", "is_cycling", "duration_sec", "distance_mi",
    "tss", "work_kj", "np_w", "avg_hr_bpm", "max_hr_bpm", "cadence_rpm", "if_", "ef", "vi",
    "p5s_w", "p1min_w", "p5min_w", "p10min_w", "p20min_w", "p1hr_w", "p2hr_w", "rpe",
    "feeling", "anaerobic_tis", "aerobic_tis", "pwhr_pct", "source_file",
]
_DAILY_COLS = [
    "date", "year", "is_projected", "has_ride", "num_workouts", "tss_sum", "duration_sec",
    "distance_mi", "work_kj", "if_daily", "atl", "ctl", "tsb", "mftp_w", "frc_kj", "pmax_w",
    "tte_sec", "weight_lb", "fat_pct", "sickness", "hrv_7d_avg_ms", "hrv_daily_ms", "rhr_bpm",
    "sleep_total_sec", "sleep_deep_sec", "sleep_light_sec", "sleep_rem_sec", "sleep_awake_sec",
    "tiz_pwr_z1_sec", "tiz_pwr_z2_sec", "tiz_pwr_z3_sec", "tiz_pwr_z4_sec", "tiz_pwr_z5_sec",
    "tiz_pwr_z6_sec", "tiz_hr_z1_sec", "tiz_hr_z2_sec", "tiz_hr_z3_sec", "tiz_hr_z4_sec",
    "tiz_hr_z5_sec", "data_flags",
]


def _insert_workouts(conn, workouts):
    ph = ",".join("?" * len(_WORKOUT_COLS))
    conn.executemany(
        f"INSERT INTO workout ({','.join(_WORKOUT_COLS)}) VALUES ({ph})",
        [tuple(w.get(c) for c in _WORKOUT_COLS) for w in workouts],
    )


def _insert_daily(conn, rows):
    ph = ",".join("?" * len(_DAILY_COLS))
    conn.executemany(
        f"INSERT INTO daily ({','.join(_DAILY_COLS)}) VALUES ({ph})",
        [tuple(r.get(c) for c in _DAILY_COLS) for r in rows],
    )
