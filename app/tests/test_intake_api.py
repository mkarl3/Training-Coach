"""Intake backend gaps (handoff 2) — API behavior on a COLD START (no wko.db). Points the app's
DB paths at temp files so the real coach.db is untouched, then drives the endpoints via TestClient.
The findings-modifier logic + first-read voice are covered by slice2 tests / live verification;
here we pin the cold-start tolerance, the validations, and the CRUD/coercion wiring."""
import os

import pytest
from fastapi.testclient import TestClient


@pytest.fixture
def cold(tmp_path, monkeypatch, app_main):
    main = app_main
    monkeypatch.setattr(main, "WKO_DB", str(tmp_path / "absent.db"))    # missing -> cold start
    monkeypatch.setattr(main, "COACH_DB", str(tmp_path / "coach.db"))   # temp; real one untouched
    monkeypatch.setattr(main, "INDEX_DB", str(tmp_path / "idx.db"))
    main._S.clear()
    with TestClient(main.app) as client:                               # triggers startup
        yield client


# checkpoint 1 — cold start boots, status is right, nothing crashes
def test_cold_start_boots_and_status_is_falsey(cold):
    assert cold.get("/api/health").json()["loaded"] is True
    s = cold.get("/api/intake/status").json()
    assert s == {"has_data": False, "months_of_history": 0.0, "has_profile": False,
                 "has_season_or_goal": False, "complete": False}


def test_meta_and_plan_do_not_crash_when_empty(cold):
    assert cold.get("/api/meta").json()["board_status"] == "awaiting"
    assert "error" in cold.get("/api/plan").json()


# checkpoint 4 — general_goal persists + reaches read-through; invalid -> 400; no plan generated
def test_general_goal_persists_invalid_rejected(cold):
    bad = cold.post("/api/season", json={"name": "S", "start_date": "2026-06-01",
                                         "weekly_hours_budget": 7, "general_goal": "threshold"})
    assert bad.status_code == 400                                       # not one of the 4 allowed
    ok = cold.post("/api/season", json={"name": "S", "start_date": "2026-06-01",
                                        "weekly_hours_budget": 7, "general_goal": "durability"})
    assert ok.status_code == 200 and ok.json()["plan"] is None         # no A-race -> no plan
    assert cold.get("/api/season").json()["season"]["general_goal"] == "durability"
    st = cold.get("/api/intake/status").json()
    assert st["has_season_or_goal"] is True                            # general_goal counts


# checkpoint 3 (wiring) — life-event CRUD + validation; board re-applies (effect logic in slice2)
def test_life_event_validation_and_crud(cold):
    assert cold.post("/api/life-event", json={"start_date": "2026-03-01", "category": "sickness"}).status_code == 400
    assert cold.post("/api/life-event", json={"start_date": "2026-03-01", "category": "injury",
                                              "detector_effect": "nuke"}).status_code == 400
    r = cold.post("/api/life-event", json={"start_date": "2026-03-01", "end_date": "2026-03-20",
                                           "category": "injury"})
    assert r.status_code == 200                                        # default effect applied, no crash
    evs = cold.get("/api/life-event").json()["life_events"]
    assert len(evs) == 1 and evs[0]["detector_effect"] == "downgrade_severity"
    assert cold.delete(f"/api/life-event/{evs[0]['id']}").status_code == 200
    assert cold.get("/api/life-event").json()["life_events"] == []


# checkpoint 5 — weight_kg empty -> None; a value round-trips
def test_weight_kg_coercion(cold):
    cold.post("/api/profile", json={"updates": {"weight_kg": 72.5}})
    assert cold.get("/api/profile").json()["profile"]["weight_kg"] == 72.5
    cold.post("/api/profile", json={"updates": {"weight_kg": ""}})
    assert cold.get("/api/profile").json()["profile"]["weight_kg"] is None


# endpoints that need data return 409 (not 500) while awaiting intake
def test_data_endpoints_409_when_awaiting(cold):
    assert cold.get("/api/watchman?as_of=2026-01-01").status_code == 409
    assert cold.post("/api/coach/message", json={"text": "hi"}).status_code == 409
    assert cold.post("/api/coach/first-read").status_code == 409


# checkpoint 2 — intake-mode upload of <12 months is rejected (422), live dataset untouched.
def _xlsx_bytes():
    # a minimal WKO-SHAPED workbook so content classification accepts it (row 2 = header names
    # with a recognized PMC column, row 3 = units with the 'date' marker). Filename is irrelevant.
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "PMC", "report"          # row 1 title (ignored by header_map)
    ws["A2"], ws["B2"] = "Date", "CTL"            # row 2 header names -> classify_sheet -> PMC
    ws["A3"], ws["B3"] = "date", "TSS/d"          # row 3 units -> date column detected
    b = io.BytesIO()
    wb.save(b)
    return b.getvalue()


def _short_db(days):
    """A fake builder: writes a tmp db whose daily span is `days` long (is_projected=0)."""
    import datetime
    import sqlite3

    def build(tmp, exports, loaded_at=None):
        c = sqlite3.connect(tmp)
        c.execute("CREATE TABLE daily (date TEXT, is_projected INT)")
        base = datetime.date(2025, 1, 1)
        for i in range(days + 1):
            c.execute("INSERT INTO daily VALUES (?,0)", ((base + datetime.timedelta(days=i)).isoformat(),))
        c.commit()
        c.close()
    return build


def test_intake_upload_rejects_under_12_months(cold, app_main, monkeypatch, tmp_path):
    main = app_main
    monkeypatch.setattr(main, "EXPORTS_DIR", str(tmp_path / "exports"))   # don't touch real exports
    monkeypatch.setattr(main.loader, "build_database", _short_db(100))    # ~100-day span
    monkeypatch.setattr(main.validator, "run", lambda *a, **k: {"round_trip_ok": True, "round_trip": []})
    files = {"files": ("Training History test.xlsx", _xlsx_bytes(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = cold.post("/api/upload?intake=true", files=files)
    assert r.status_code == 422 and "12 months" in r.json()["detail"]
    assert not os.path.exists(main.WKO_DB)                                # live dataset untouched


def test_incremental_upload_has_no_minimum(cold, app_main, monkeypatch, tmp_path):
    # same short file with intake=False passes the gate (no minimum) — it only fails later when the
    # fake short db can't hot-reload as real metrics, which is past the gate we're testing.
    main = app_main
    monkeypatch.setattr(main, "EXPORTS_DIR", str(tmp_path / "exports"))
    monkeypatch.setattr(main.loader, "build_database", _short_db(100))
    monkeypatch.setattr(main.validator, "run", lambda *a, **k: {"round_trip_ok": True, "round_trip": []})
    monkeypatch.setattr(main, "_load_training_data", lambda: None)   # isolate the gate from hot-reload
    files = {"files": ("Training History test.xlsx", _xlsx_bytes(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = cold.post("/api/upload", files=files)            # intake defaults False
    assert r.status_code == 200 and "12 months" not in r.text        # the 12-month gate did NOT fire


def test_multi_file_upload_rolls_back_whole_batch(cold, app_main, monkeypatch, tmp_path):
    # several files in one request; if the rebuild is rejected, the WHOLE batch is rolled back.
    main = app_main
    exports = tmp_path / "exports"
    monkeypatch.setattr(main, "EXPORTS_DIR", str(exports))
    monkeypatch.setattr(main.loader, "build_database", _short_db(100))   # <12mo -> intake 422
    monkeypatch.setattr(main.validator, "run", lambda *a, **k: {"round_trip_ok": True, "round_trip": []})
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    files = [("files", ("Training History 2024.xlsx", _xlsx_bytes(), mime)),
             ("files", ("PMC Report 2024.xlsx", _xlsx_bytes(), mime))]
    r = cold.post("/api/upload?intake=true", files=files)
    assert r.status_code == 422
    assert not (exports / "Training History 2024.xlsx").exists()        # both rolled back
    assert not (exports / "PMC Report 2024.xlsx").exists()


def test_upload_classifies_by_content_not_filename(cold, app_main, monkeypatch, tmp_path):
    import io
    import openpyxl
    main = app_main
    monkeypatch.setattr(main, "EXPORTS_DIR", str(tmp_path / "exports"))
    monkeypatch.setattr(main.loader, "build_database", _short_db(400))   # >12 months -> gate passes
    monkeypatch.setattr(main.validator, "run", lambda *a, **k: {"round_trip_ok": True, "round_trip": []})
    monkeypatch.setattr(main, "_load_training_data", lambda: None)
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # arbitrary filename + real WKO content -> accepted
    ok = cold.post("/api/upload?intake=true", files={"files": ("2025.xlsx", _xlsx_bytes(), mime)})
    assert ok.status_code == 200
    # a readable .xlsx that ISN'T a WKO export -> rejected on CONTENT (not the filename)
    wb = openpyxl.Workbook(); wb.active["A1"] = "hello"; b = io.BytesIO(); wb.save(b)
    bad = cold.post("/api/upload?intake=true", files={"files": ("Training History 2025.xlsx", b.getvalue(), mime)})
    assert bad.status_code == 400 and "WKO5 export" in bad.json()["detail"]
